import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.max_row

def get_column_count(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.max_column

def read_data(file,sheetName,row_num, column_no):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(row_num,column_no).value

def write_data(file, sheetName,row_num,column_no,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(row_num, column_no).value=data
    workbook.save(file)

def fill_green_color(file,sheetName, rownum, columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    greenfill=PatternFill(start_color='60b212',
                        end_color='60b212',
                        fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenfill
    workbook.save(file)

def fill_red_color(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redfill = PatternFill(start_color='ff0000',
                                end_color='ff0000',
                                fill_type='solid')
    sheet.cell(rownum, columnno).fill = redfill
    workbook.save(file)

